import openpyxl as xl
from openpyxl.chart import BarChart, Reference
from pathlib import Path


def process_sheets(file):
    wb = xl.load_workbook(file)
    sheet = wb["Sheet1"]

    rows = sheet.max_row
    sheet.cell(1,4).value = 'Sales Price'

    for row in range(2, rows + 1):
        discounted_val = sheet.cell(row, 3).value * 0.9
        sheet.cell(row, 4).value = discounted_val

    values_of_new_prices = Reference(sheet, 2, sheet.max_row, 4, 4)

    chart = BarChart()
    chart.add_data(values_of_new_prices)

    sheet.add_chart(chart, "g1")

    wb.save(file)


path = Path()
for filename in path.glob('*.xlsx'):
    process_sheets(filename)











